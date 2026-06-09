import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import OutlineProperties

from src.utils.taxonomy import GOVERNANCE, TIER1, TIER2

# ── fill palette ──────────────────────────────────────────────────────────────
_YES_FILL         = PatternFill("solid", fgColor="BBDEFB")  # blue  (Yes)
_NO_FILL          = PatternFill("solid", fgColor="F5F5F5")  # grey  (No)
_HEAD_FILL        = PatternFill("solid", fgColor="D1D5DB")  # grey  header
_ADOPTED_HDR_FILL = PatternFill("solid", fgColor="E8F5E9")  # mint  adopted header

_STAT_HIGH  = PatternFill("solid", fgColor="A5D6A7")  # strong green
_STAT_MED   = PatternFill("solid", fgColor="C8E6C9")  # light  green
_STAT_LOW   = PatternFill("solid", fgColor="F5F5F5")  # near-white

_MISS_HIGH  = PatternFill("solid", fgColor="FFCDD2")  # red   (bad)
_MISS_MED   = PatternFill("solid", fgColor="FFECB3")  # amber

_T1_FILLS = {
    "S1":  PatternFill("solid", fgColor="DBEAFE"),
    "S2":  PatternFill("solid", fgColor="E0E7FF"),
    "S3U": PatternFill("solid", fgColor="DCFCE7"),
    "S3D": PatternFill("solid", fgColor="D1FAE5"),
    "CDR": PatternFill("solid", fgColor="FEF9C3"),
}
_CAT_FILLS = {
    "Tier 1":     PatternFill("solid", fgColor="DBEAFE"),
    "Tier 2":     PatternFill("solid", fgColor="DCFCE7"),
    "Governance": PatternFill("solid", fgColor="FEF9C3"),
}
_SUMMARY_FILL = PatternFill("solid", fgColor="EEEEEE")


def _is_true(v) -> bool:
    return v is True or str(v).strip().lower() == "true"


def _adopted_display(v) -> str:
    if _is_true(v):
        return "Yes"
    if v is False or str(v).strip().lower() == "false":
        return "No"
    return ""


def _chunk_flags(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(False, index=df.index)
    return df[col].apply(_is_true)


def _pct_fill(pct: float) -> PatternFill:
    if pct >= 50:
        return _STAT_HIGH
    if pct >= 20:
        return _STAT_MED
    return _STAT_LOW


def _miss_fill(pct: float) -> PatternFill:
    if pct >= 50:
        return _MISS_HIGH
    if pct >= 20:
        return _MISS_MED
    return _STAT_LOW


def _set_col_widths(ws, headers: list[str], widths: dict[str, float], default: float = 14):
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, default)


class ExcelReporter:
    def __init__(self, df: pd.DataFrame):
        self.df = df

    # ── public entry point ────────────────────────────────────────────────────

    def write(self, path: str):
        df    = self.df
        id_cols = ["companyname", "filingDate", "chunk_ids"]

        tier1_cols = [f"tier1_{s}_{sfx}" for s in TIER1 for sfx in ("adopted", "quote")]
        tier2_cols = {
            t1: [f"tier2_{s}_{sfx}" for s in subs for sfx in ("adopted", "quote")]
            for t1, subs in TIER2.items()
        }
        gov_cols = [f"governance_{s}_{sfx}" for s in GOVERNANCE for sfx in ("adopted", "quote")]

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # ── chunk-level data sheets ──
            self._write_data_sheet(writer, "Tier 1",    df, id_cols + ["notes_tier1"], tier1_cols)
            for t1_key, cols in tier2_cols.items():
                self._write_data_sheet(writer, t1_key,  df, id_cols + ["notes_tier2"], cols)
            self._write_data_sheet(writer, "Governance", df, id_cols, gov_cols)

            # ── stats sheets ──
            self._write_adoption(writer.book.create_sheet("Stats - Adoption"))
            self._write_coverage(writer.book.create_sheet("Stats - Coverage"))
            self._write_t1_accuracy(writer.book.create_sheet("Stats - T1 Accuracy"))

            # ── raw dump ──
            df.to_excel(writer, index=False, sheet_name="Raw")

            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]

    # ── data sheet (chunk-level) ──────────────────────────────────────────────

    def _write_data_sheet(self, writer, sheet_name: str, df: pd.DataFrame,
                          id_cols: list[str], strategy_cols: list[str]):
        ws = writer.book.create_sheet(sheet_name)
        all_cols = [c for c in id_cols if c in df.columns] + \
                   [c for c in strategy_cols if c in df.columns]
        sub = df[all_cols].copy()

        adopted_set = {c for c in strategy_cols if c.endswith("_adopted") and c in sub.columns}
        for col in adopted_set:
            sub[col] = sub[col].apply(_adopted_display)

        bold   = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        wrap   = Alignment(wrap_text=True, vertical="top")

        for ci, col in enumerate(all_cols, 1):
            cell = ws.cell(1, ci, col)
            cell.font  = bold
            cell.fill  = _ADOPTED_HDR_FILL if col in adopted_set else _HEAD_FILL
            cell.alignment = center

        for ri, (_, row) in enumerate(sub.iterrows(), 2):
            for ci, col in enumerate(all_cols, 1):
                val  = row[col]
                cell = ws.cell(ri, ci, val if val != "" else None)
                if col in adopted_set:
                    cell.fill      = _YES_FILL if val == "Yes" else (_NO_FILL if val == "No" else PatternFill())
                    cell.alignment = center
                elif col.endswith("_quote") or col in ("chunks", "notes_tier1", "notes_tier2"):
                    cell.alignment = wrap

        n_id = len([c for c in id_cols if c in df.columns])
        for ci, col in enumerate(all_cols, 1):
            letter = get_column_letter(ci)
            if col in adopted_set:
                ws.column_dimensions[letter].width = 8
            elif col.endswith("_quote") or col in ("chunks",):
                ws.column_dimensions[letter].width = 40
            elif col in ("notes_tier1", "notes_tier2"):
                ws.column_dimensions[letter].width = 35
            else:
                ws.column_dimensions[letter].width = max(len(col) + 2, 14)

        # group quote + notes columns so they can be collapsed with the +/- button
        for ci, col in enumerate(all_cols, 1):
            if col.endswith("_quote") or col in ("notes_tier1", "notes_tier2", "chunks"):
                ws.column_dimensions[get_column_letter(ci)].outlineLevel = 1
        # place the collapse button to the left of the group (more intuitive)
        ws.sheet_properties.outlinePr = OutlineProperties(summaryRight=False)

        ws.freeze_panes    = ws.cell(2, n_id + 1)
        ws.row_dimensions[1].height = 20

    # ── stats: adoption rate ──────────────────────────────────────────────────

    def _build_adoption_stats(self) -> pd.DataFrame:
        df, n = self.df, len(self.df)
        rows = []
        for s in TIER1:
            f = _chunk_flags(df, f"tier1_{s}_adopted")
            rows.append({"Category": "Tier 1", "Group": "", "Strategy": s,
                         "Chunks True": int(f.sum()), "Total Chunks": n,
                         "% True": round(f.mean() * 100, 1)})
        for t1, subs in TIER2.items():
            for s in subs:
                f = _chunk_flags(df, f"tier2_{s}_adopted")
                rows.append({"Category": "Tier 2", "Group": t1, "Strategy": s,
                             "Chunks True": int(f.sum()), "Total Chunks": n,
                             "% True": round(f.mean() * 100, 1)})
        for s in GOVERNANCE:
            f = _chunk_flags(df, f"governance_{s}_adopted")
            rows.append({"Category": "Governance", "Group": "", "Strategy": s,
                         "Chunks True": int(f.sum()), "Total Chunks": n,
                         "% True": round(f.mean() * 100, 1)})
        return pd.DataFrame(rows)

    def _write_adoption(self, ws):
        stats  = self._build_adoption_stats()
        bold   = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        headers = list(stats.columns)

        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = bold; cell.fill = _HEAD_FILL; cell.alignment = center

        for r, row in enumerate(stats.itertuples(index=False), 2):
            for c, val in enumerate(row, 1):
                col = headers[c - 1]
                cell = ws.cell(r, c, val)
                if col == "% True":
                    cell.fill = _pct_fill(val); cell.alignment = center
                elif col in ("Chunks True", "Total Chunks"):
                    cell.alignment = center
                elif col == "Category":
                    cell.fill = _CAT_FILLS.get(val, PatternFill())

        _set_col_widths(ws, headers,
                        {"Category": 14, "Group": 10, "Strategy": 28,
                         "Chunks True": 14, "Total Chunks": 14, "% True": 10})
        ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 20

    # ── stats: T1 → T2 coverage ───────────────────────────────────────────────

    def _build_coverage_stats(self) -> pd.DataFrame:
        df   = self.df
        rows = []
        for t1, subs in TIER2.items():
            t1f   = _chunk_flags(df, f"tier1_{t1}_adopted")
            n_t1  = int(t1f.sum())
            any_t2 = pd.Series(False, index=df.index)
            for s in subs:
                t2f    = _chunk_flags(df, f"tier2_{s}_adopted")
                n_both = int((t1f & t2f).sum())
                rows.append({"Tier 1": t1, "Tier 2 Strategy": s,
                             "Chunks (T1=True)": n_t1,
                             "Chunks (Both True)": n_both,
                             "Coverage %": round(n_both / n_t1 * 100, 1) if n_t1 else 0.0,
                             "Unpinpointed": ""})
                any_t2 = any_t2 | t2f
            n_unpin = int((t1f & ~any_t2).sum())
            rows.append({"Tier 1": t1, "Tier 2 Strategy": "— no T2 match —",
                         "Chunks (T1=True)": n_t1, "Chunks (Both True)": "",
                         "Coverage %": "",
                         "Unpinpointed": round(n_unpin / n_t1 * 100, 1) if n_t1 else 0.0})
        return pd.DataFrame(rows)

    def _write_coverage(self, ws):
        cov    = self._build_coverage_stats()
        bold   = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        headers = list(cov.columns)

        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = bold; cell.fill = _HEAD_FILL; cell.alignment = center

        for r, row in enumerate(cov.itertuples(index=False), 2):
            t1, is_sum = row[0], row[1] == "— no T2 match —"
            for c, val in enumerate(row, 1):
                col  = headers[c - 1]
                cell = ws.cell(r, c, val if val != "" else None)
                if is_sum:
                    cell.font = Font(bold=True, italic=True)
                    if col != "Unpinpointed":
                        cell.fill = _SUMMARY_FILL
                if col == "Coverage %" and not is_sum:
                    cell.fill = _pct_fill(val if isinstance(val, (int, float)) else 0)
                    cell.alignment = center
                elif col == "Unpinpointed" and is_sum:
                    cell.fill = _miss_fill(val if isinstance(val, (int, float)) else 0)
                    cell.alignment = center
                elif col.startswith("Chunks"):
                    cell.alignment = center
                elif col == "Tier 1":
                    cell.fill = _T1_FILLS.get(t1, PatternFill())

        _set_col_widths(ws, headers,
                        {"Tier 1": 10, "Tier 2 Strategy": 22, "Chunks (T1=True)": 16,
                         "Chunks (Both True)": 18, "Coverage %": 13, "Unpinpointed": 15})
        ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 20

    # ── stats: T2 → T1 accuracy ───────────────────────────────────────────────

    def _build_t1_accuracy_stats(self) -> pd.DataFrame:
        df   = self.df
        rows = []
        for t1, subs in TIER2.items():
            t1f = _chunk_flags(df, f"tier1_{t1}_adopted")
            for s in subs:
                t2f      = _chunk_flags(df, f"tier2_{s}_adopted")
                n_t2     = int(t2f.sum())
                n_orphan = int((t2f & ~t1f).sum())
                rows.append({"Tier 1": t1, "Tier 2 Strategy": s,
                             "Chunks (T2=True)": n_t2,
                             "T2 True, T1 False": n_orphan,
                             "T1 Miss %": round(n_orphan / n_t2 * 100, 1) if n_t2 else 0.0})
        return pd.DataFrame(rows)

    def _write_t1_accuracy(self, ws):
        acc    = self._build_t1_accuracy_stats()
        bold   = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        headers = list(acc.columns)

        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = bold; cell.fill = _HEAD_FILL; cell.alignment = center

        for r, row in enumerate(acc.itertuples(index=False), 2):
            t1 = row[0]
            for c, val in enumerate(row, 1):
                col  = headers[c - 1]
                cell = ws.cell(r, c, val)
                if col == "T1 Miss %":
                    cell.fill = _miss_fill(val if isinstance(val, (int, float)) else 0)
                    cell.alignment = center
                elif col in ("Chunks (T2=True)", "T2 True, T1 False"):
                    cell.alignment = center
                elif col == "Tier 1":
                    cell.fill = _T1_FILLS.get(t1, PatternFill())

        _set_col_widths(ws, headers,
                        {"Tier 1": 10, "Tier 2 Strategy": 25,
                         "Chunks (T2=True)": 16, "T2 True, T1 False": 18, "T1 Miss %": 13})
        ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 20
