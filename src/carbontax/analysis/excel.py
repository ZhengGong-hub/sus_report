"""
excel.py — render the combined parsed output into a styled multi-sheet workbook.

Input is the flat CSV from extraction.parse_output (one row per chunk, bare-boolean
flags plus per-measure _quote/_page evidence). Output mirrors the legacy review
workbook:

  Tier 1            — chunk-level S1/S2/S3U/S3D/CDR adoption (Yes/No)
  S1 … CDR          — one sheet per tier-1 bucket, its tier-2 measures with quote+page
  Governance        — sbti / icp / exec-comp / assurance with quote+page
  Stats - Adoption  — % of chunks flagged True, per strategy
  Stats - Coverage  — of chunks where a tier-1 bucket is True, how many also pin a
                      tier-2 measure (and the unpinpointed remainder)
  Stats - T1 Accuracy — of chunks where a tier-2 measure is True, how many have the
                      parent tier-1 bucket False (a tier-1 miss)
  Raw               — full flat dump

This is the combined-schema successor to legacy/excel.py: flags are bare booleans
(tier1_S1, not tier1_S1_adopted), evidence carries a page number, and there are no
per-call notes columns.

Usage (run from repo root):
  carbontax-report --run-name pilot
  carbontax-report --run-name pilot --input <parsed.csv> --dest <out.xlsx>
"""

from __future__ import annotations

import argparse

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

from carbontax.extraction.paths import DEFAULT_RUN_NAME, parsed_csv, report_xlsx
from carbontax.taxonomy import GOVERNANCE_FLAGS, MEASURE_IDS, MEASURE_SCOPE, TIER1_BUCKETS

# tier-1 buckets in canonical order, and the tier-2 measures grouped beneath each
# (MEASURE_IDS is already ordered by scope, so this preserves the intended order).
TIER1 = list(TIER1_BUCKETS)
TIER2: dict[str, list[str]] = {b: [] for b in TIER1}
for _mid in MEASURE_IDS:
    TIER2[MEASURE_SCOPE[_mid]].append(_mid)
GOVERNANCE = list(GOVERNANCE_FLAGS)

# ── fill palette ──────────────────────────────────────────────────────────────
_YES_FILL         = PatternFill("solid", fgColor="BBDEFB")  # blue  (Yes)
_NO_FILL          = PatternFill("solid", fgColor="F5F5F5")  # grey  (No)
_HEAD_FILL        = PatternFill("solid", fgColor="D1D5DB")  # grey  header
_ADOPTED_HDR_FILL = PatternFill("solid", fgColor="E8F5E9")  # mint  adopted header

_STAT_HIGH  = PatternFill("solid", fgColor="A5D6A7")  # strong green
_STAT_MED   = PatternFill("solid", fgColor="C8E6C9")  # light  green
_STAT_LOW   = PatternFill("solid", fgColor="F5F5F5")  # near-white

_MISS_HIGH  = PatternFill("solid", fgColor="FFCDD2")  # red   (bad)
_MISS_MED   = PatternFill("solid", fgColor="FFECB3")  # amber

_T1_FILLS = {
    "S1":  PatternFill("solid", fgColor="DBEAFE"),
    "S2":  PatternFill("solid", fgColor="E0E7FF"),
    "S3U": PatternFill("solid", fgColor="DCFCE7"),
    "S3D": PatternFill("solid", fgColor="D1FAE5"),
    "CDR": PatternFill("solid", fgColor="FEF9C3"),
}
_CAT_FILLS = {
    "Tier 1":     PatternFill("solid", fgColor="DBEAFE"),
    "Tier 2":     PatternFill("solid", fgColor="DCFCE7"),
    "Governance": PatternFill("solid", fgColor="FEF9C3"),
}
_SUMMARY_FILL = PatternFill("solid", fgColor="EEEEEE")

_BOLD   = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _is_true(v) -> bool:
    return v is True or str(v).strip().lower() == "true"


def _adopted_display(v) -> str:
    if _is_true(v):
        return "Yes"
    if v is False or str(v).strip().lower() == "false":
        return "No"
    return ""


def _chunk_flags(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(False, index=df.index)
    return df[col].apply(_is_true)


def _pct_fill(pct: float) -> PatternFill:
    if pct >= 50:
        return _STAT_HIGH
    if pct >= 20:
        return _STAT_MED
    return _STAT_LOW


def _miss_fill(pct: float) -> PatternFill:
    if pct >= 50:
        return _MISS_HIGH
    if pct >= 20:
        return _MISS_MED
    return _STAT_LOW


def _set_col_widths(ws, headers: list[str], widths: dict[str, float], default: float = 14):
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, default)


class ExcelReporter:
    """Render a combined-schema parsed DataFrame into a styled workbook."""

    def __init__(self, df: pd.DataFrame):
        self.df = df

    # ── public entry point ────────────────────────────────────────────────────

    def write(self, path: str):
        df = self.df
        id_cols = ["companyname", "filingDate", "chunk_ids"]

        # tier1: bare boolean flag per bucket (no quote/page in the combined schema)
        tier1_cols = [f"tier1_{b}" for b in TIER1]
        # tier2: flag + quote + page per measure, grouped by tier-1 bucket
        tier2_cols = {
            b: [c for m in subs for c in (f"tier2_{m}", f"tier2_{m}_quote", f"tier2_{m}_page")]
            for b, subs in TIER2.items()
        }
        gov_cols = [
            c for f in GOVERNANCE
            for c in (f"governance_{f}", f"governance_{f}_quote", f"governance_{f}_page")
        ]

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # ── chunk-level data sheets ──
            self._write_data_sheet(writer, "Tier 1", df, id_cols, tier1_cols)
            for t1_key, cols in tier2_cols.items():
                self._write_data_sheet(writer, t1_key, df, id_cols, cols)
            self._write_data_sheet(writer, "Governance", df, id_cols, gov_cols)

            # ── stats sheets ──
            self._write_adoption(writer.book.create_sheet("Stats - Adoption"))
            self._write_coverage(writer.book.create_sheet("Stats - Coverage"))
            self._write_t1_accuracy(writer.book.create_sheet("Stats - T1 Accuracy"))

            # ── raw dump ──
            df.to_excel(writer, index=False, sheet_name="Raw")

            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]

    # ── data sheet (chunk-level) ──────────────────────────────────────────────

    def _write_data_sheet(self, writer, sheet_name: str, df: pd.DataFrame,
                          id_cols: list[str], strategy_cols: list[str]):
        ws = writer.book.create_sheet(sheet_name)
        all_cols = [c for c in id_cols if c in df.columns] + \
                   [c for c in strategy_cols if c in df.columns]
        sub = df[all_cols].copy()

        # adopted = the bare flag columns (not the _quote / _page evidence columns)
        adopted_set = {
            c for c in strategy_cols
            if c in sub.columns and not c.endswith(("_quote", "_page"))
        }
        for col in adopted_set:
            sub[col] = sub[col].apply(_adopted_display)

        wrap = Alignment(wrap_text=True, vertical="top")

        for ci, col in enumerate(all_cols, 1):
            cell = ws.cell(1, ci, col)
            cell.font  = _BOLD
            cell.fill  = _ADOPTED_HDR_FILL if col in adopted_set else _HEAD_FILL
            cell.alignment = _CENTER

        for ri, (_, row) in enumerate(sub.iterrows(), 2):
            for ci, col in enumerate(all_cols, 1):
                val = row[col]
                blank = val == "" or pd.isna(val)
                cell = ws.cell(ri, ci, None if blank else val)
                if col in adopted_set:
                    cell.fill      = _YES_FILL if val == "Yes" else (_NO_FILL if val == "No" else PatternFill())
                    cell.alignment = _CENTER
                elif col.endswith("_page"):
                    cell.alignment = _CENTER
                elif col.endswith("_quote") or col == "chunks":
                    cell.alignment = wrap

        n_id = len([c for c in id_cols if c in df.columns])
        for ci, col in enumerate(all_cols, 1):
            letter = get_column_letter(ci)
            if col in adopted_set:
                ws.column_dimensions[letter].width = 8
            elif col.endswith("_page"):
                ws.column_dimensions[letter].width = 7
            elif col.endswith("_quote") or col == "chunks":
                ws.column_dimensions[letter].width = 40
            else:
                ws.column_dimensions[letter].width = max(len(col) + 2, 14)

        # group quote/page evidence columns so they collapse with the +/- button
        for ci, col in enumerate(all_cols, 1):
            if col.endswith(("_quote", "_page")) or col == "chunks":
                ws.column_dimensions[get_column_letter(ci)].outlineLevel = 1
        # place the collapse button to the left of the group (more intuitive)
        ws.sheet_properties.outlinePr = Outline(summaryRight=False)

        ws.freeze_panes    = ws.cell(2, n_id + 1)
        ws.row_dimensions[1].height = 20

    # ── generic stats-sheet writer ────────────────────────────────────────────

    def _write_stats_sheet(self, ws, stats: pd.DataFrame, widths: dict[str, float],
                           style_cell) -> None:
        """Write a stats DataFrame with a styled header row.

        ``style_cell(cell, col, val, row)`` applies per-cell fill/font/alignment;
        ``row`` is the full namedtuple so styles can depend on sibling columns.
        """
        headers = list(stats.columns)

        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = _BOLD
            cell.fill = _HEAD_FILL
            cell.alignment = _CENTER

        for r, row in enumerate(stats.itertuples(index=False), 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(r, c, None if val == "" else val)
                style_cell(cell, headers[c - 1], val, row)

        _set_col_widths(ws, headers, widths)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

    # ── stats: adoption rate ──────────────────────────────────────────────────

    def _build_adoption_stats(self) -> pd.DataFrame:
        df, n = self.df, len(self.df)
        rows = []
        for s in TIER1:
            f = _chunk_flags(df, f"tier1_{s}")
            rows.append({"Category": "Tier 1", "Group": "", "Strategy": s,
                         "Chunks True": int(f.sum()), "Total Chunks": n,
                         "% True": round(f.mean() * 100, 1)})
        for t1, subs in TIER2.items():
            for s in subs:
                f = _chunk_flags(df, f"tier2_{s}")
                rows.append({"Category": "Tier 2", "Group": t1, "Strategy": s,
                             "Chunks True": int(f.sum()), "Total Chunks": n,
                             "% True": round(f.mean() * 100, 1)})
        for s in GOVERNANCE:
            f = _chunk_flags(df, f"governance_{s}")
            rows.append({"Category": "Governance", "Group": "", "Strategy": s,
                         "Chunks True": int(f.sum()), "Total Chunks": n,
                         "% True": round(f.mean() * 100, 1)})
        return pd.DataFrame(rows)

    def _write_adoption(self, ws):
        def style_cell(cell, col, val, row):
            if col == "% True":
                cell.fill = _pct_fill(val)
                cell.alignment = _CENTER
            elif col in ("Chunks True", "Total Chunks"):
                cell.alignment = _CENTER
            elif col == "Category":
                cell.fill = _CAT_FILLS.get(val, PatternFill())

        self._write_stats_sheet(
            ws, self._build_adoption_stats(),
            widths={"Category": 14, "Group": 10, "Strategy": 28,
                    "Chunks True": 14, "Total Chunks": 14, "% True": 10},
            style_cell=style_cell,
        )

    # ── stats: T1 → T2 coverage ───────────────────────────────────────────────

    def _build_coverage_stats(self) -> pd.DataFrame:
        df   = self.df
        rows = []
        for t1, subs in TIER2.items():
            t1f   = _chunk_flags(df, f"tier1_{t1}")
            n_t1  = int(t1f.sum())
            any_t2 = pd.Series(False, index=df.index)
            for s in subs:
                t2f    = _chunk_flags(df, f"tier2_{s}")
                n_both = int((t1f & t2f).sum())
                rows.append({"Tier 1": t1, "Tier 2 Strategy": s,
                             "Chunks (T1=True)": n_t1,
                             "Chunks (Both True)": n_both,
                             "Coverage %": round(n_both / n_t1 * 100, 1) if n_t1 else 0.0,
                             "Unpinpointed": ""})
                any_t2 = any_t2 | t2f
            n_unpin = int((t1f & ~any_t2).sum())
            rows.append({"Tier 1": t1, "Tier 2 Strategy": "— no T2 match —",
                         "Chunks (T1=True)": n_t1, "Chunks (Both True)": "",
                         "Coverage %": "",
                         "Unpinpointed": round(n_unpin / n_t1 * 100, 1) if n_t1 else 0.0})
        return pd.DataFrame(rows)

    def _write_coverage(self, ws):
        def style_cell(cell, col, val, row):
            t1, is_sum = row[0], row[1] == "— no T2 match —"
            if is_sum:
                cell.font = Font(bold=True, italic=True)
                if col != "Unpinpointed":
                    cell.fill = _SUMMARY_FILL
            if col == "Coverage %" and not is_sum:
                cell.fill = _pct_fill(val if isinstance(val, (int, float)) else 0)
                cell.alignment = _CENTER
            elif col == "Unpinpointed" and is_sum:
                cell.fill = _miss_fill(val if isinstance(val, (int, float)) else 0)
                cell.alignment = _CENTER
            elif col.startswith("Chunks"):
                cell.alignment = _CENTER
            elif col == "Tier 1":
                cell.fill = _T1_FILLS.get(t1, PatternFill())

        self._write_stats_sheet(
            ws, self._build_coverage_stats(),
            widths={"Tier 1": 10, "Tier 2 Strategy": 22, "Chunks (T1=True)": 16,
                    "Chunks (Both True)": 18, "Coverage %": 13, "Unpinpointed": 15},
            style_cell=style_cell,
        )

    # ── stats: T2 → T1 accuracy ───────────────────────────────────────────────

    def _build_t1_accuracy_stats(self) -> pd.DataFrame:
        df   = self.df
        rows = []
        for t1, subs in TIER2.items():
            t1f = _chunk_flags(df, f"tier1_{t1}")
            for s in subs:
                t2f      = _chunk_flags(df, f"tier2_{s}")
                n_t2     = int(t2f.sum())
                n_orphan = int((t2f & ~t1f).sum())
                rows.append({"Tier 1": t1, "Tier 2 Strategy": s,
                             "Chunks (T2=True)": n_t2,
                             "T2 True, T1 False": n_orphan,
                             "T1 Miss %": round(n_orphan / n_t2 * 100, 1) if n_t2 else 0.0})
        return pd.DataFrame(rows)

    def _write_t1_accuracy(self, ws):
        def style_cell(cell, col, val, row):
            if col == "T1 Miss %":
                cell.fill = _miss_fill(val if isinstance(val, (int, float)) else 0)
                cell.alignment = _CENTER
            elif col in ("Chunks (T2=True)", "T2 True, T1 False"):
                cell.alignment = _CENTER
            elif col == "Tier 1":
                cell.fill = _T1_FILLS.get(row[0], PatternFill())

        self._write_stats_sheet(
            ws, self._build_t1_accuracy_stats(),
            widths={"Tier 1": 10, "Tier 2 Strategy": 25,
                    "Chunks (T2=True)": 16, "T2 True, T1 False": 18, "T1 Miss %": 13},
            style_cell=style_cell,
        )


def build_report(input_path: str, dest_path: str) -> None:
    df = pd.read_csv(input_path, dtype={"chunk_ids": "string"})
    print(f"Read {len(df)} rows from {input_path}")
    ExcelReporter(df).write(dest_path)
    print(f"Wrote workbook → {dest_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Render combined parsed output to a styled workbook")
    parser.add_argument("--run-name", default=DEFAULT_RUN_NAME, help="Run folder name under batch_folder/")
    parser.add_argument("--input", default=None, help="Parsed combined CSV (defaults to run folder)")
    parser.add_argument("--dest",  default=None, help="Destination .xlsx path (defaults to run folder)")
    args = parser.parse_args()

    build_report(
        input_path=args.input or parsed_csv(args.run_name),
        dest_path=args.dest    or report_xlsx(args.run_name),
    )


if __name__ == "__main__":
    main()
